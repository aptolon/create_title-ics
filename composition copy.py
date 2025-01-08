import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from PIL import Image, ImageTk
from docxtpl import DocxTemplate
from docx2pdf import convert
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader, PdfWriter
from io import BytesIO
import fitz
class ProjectInfoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Состав проекта")
        self.root.geometry("950x710")

        self.filepath = ""
        self.customer = tk.StringVar()
        self.project_name = tk.StringVar()
        self.project_code = tk.StringVar()  # New field
        self.ceo = tk.StringVar()  # New field
        self.project_engineer = tk.StringVar()  # New field
        self.city = tk.StringVar()  # New field
        self.year_of_development = tk.StringVar()  # New field
        self.image_path = tk.StringVar()
        self.dynamic_fields = {}  # Store dynamic fields for the second page

        # Create notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        # Create tabs
        self.tab_file_selection = ttk.Frame(self.notebook)
        self.tab_project_info = ttk.Frame(self.notebook)
        self.tab_project_contents = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_file_selection, text="Выбор файла")

        # Initially display the file selection tab
        self.create_file_selection_page()
        

    def create_file_selection_page(self):
        # Clear tab content and create content for file selection
        self.clear_tab(self.tab_file_selection)

        btn_select_file = tk.Button(self.tab_file_selection, text="Выберите файл", command=self.select_file)
        btn_select_file.place(relx=0.5, rely=0.5, anchor="center")

    def create_project_info_page(self):
        self.clear_tab(self.tab_project_info)

        # Create left and right frames
        left_frame = tk.Frame(self.tab_project_info)
        left_frame.pack(side="left", fill="both", expand=True, padx=20, pady=20)

        right_frame = tk.Frame(self.tab_project_info)
        right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)

        # Left Frame (Fields and Buttons)
        tk.Label(left_frame, text="Заказчик:").pack(anchor="w")
        customer_entry = tk.Entry(left_frame, textvariable=self.customer, width=60)
        customer_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Наименование проекта:").pack(anchor="w")
        self.project_name_entry = tk.Entry(left_frame, textvariable=self.project_name, width=60)
        self.project_name_entry.pack(anchor="w", pady=5)

        # New fields
        tk.Label(left_frame, text="Шифр проекта:").pack(anchor="w")
        project_code_entry = tk.Entry(left_frame, textvariable=self.project_code, width=60)
        project_code_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Генеральный директор:").pack(anchor="w")
        ceo_entry = tk.Entry(left_frame, textvariable=self.ceo, width=60)
        ceo_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Главный инженер проекта:").pack(anchor="w")
        project_engineer_entry = tk.Entry(left_frame, textvariable=self.project_engineer, width=60)
        project_engineer_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Город:").pack(anchor="w")
        city_entry = tk.Entry(left_frame, textvariable=self.city, width=60)
        city_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Год разработки:").pack(anchor="w")
        year_of_development_entry = tk.Entry(left_frame, textvariable=self.year_of_development, width=60)
        year_of_development_entry.pack(anchor="w", pady=5)

        tk.Label(left_frame, text="Фон:").pack(anchor="w")
        btn_select_image = tk.Button(left_frame, text="Выберите изображение", command=self.load_image_path)
        btn_select_image.pack(pady=10, anchor="w")



        btn_save = tk.Button(left_frame, text="Сохранить", command=self.save_to_excel)
        btn_save.pack(pady=10, anchor="w")
        btn_select_image = tk.Button(left_frame, text="Пример титула", command=self.overlay_text_on_image)
        btn_select_image.pack(pady=10, anchor="w")


        # Right Frame (Image Display)
        self.image_panel = tk.Label(right_frame)
        self.image_panel.pack()

        if self.image_path.get():
            self.load_image()

    def create_project_contents_page(self):
        self.clear_tab(self.tab_project_contents)

        # Create left and right frames
        left_frame = tk.Frame(self.tab_project_contents)
        left_frame.pack(side="left", fill="both", expand=True, padx=20, pady=20)

        right_frame = tk.Frame(self.tab_project_contents)
        right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)

        # Create the combobox at the top
        tk.Label(left_frame, text="Выберите из списка:").pack(anchor="w")
        self.dynamic_combobox = ttk.Combobox(left_frame, width=50)
        self.dynamic_combobox.pack(anchor="w", pady=10)
        self.dynamic_combobox.bind("<<ComboboxSelected>>", self.populate_fields_from_selection)

        # Buttons for add, delete, edit, and clear in one row
        button_frame = tk.Frame(left_frame)
        button_frame.pack(pady=10, anchor="w")

        btn_add = tk.Button(button_frame, text="Добавить", command=self.add_to_second_sheet)
        btn_add.pack(side="left", padx=5)

        btn_delete = tk.Button(button_frame, text="Удалить", command=self.delete_selected_item)
        btn_delete.pack(side="left", padx=5)

        btn_edit = tk.Button(button_frame, text="Редактировать", command=self.edit_selected_item)
        btn_edit.pack(side="left", padx=5)

        btn_clear = tk.Button(button_frame, text="Очистить поля", command=self.clear_fields)
        btn_clear.pack(side="left", padx=5)

        # Label and button for selecting a PDF file
        self.pdf_path_label = tk.Label(left_frame, text="Путь к pdf файлу:")
        self.pdf_path_label.pack(anchor="w", pady=10)

        btn_select_pdf = tk.Button(left_frame, text="Выберите pdf файл", command=self.load_pdf_file)
        btn_select_pdf.pack(pady=10, anchor="w")

        btn_create_word_documents = tk.Button(left_frame, text="Создать Word документы", command=self.create_word_documents)
        btn_create_word_documents.pack(anchor="w", pady=10)

        # Predefined fields
        fields = ["Номер тома", "Обозначение", "Наименование", "Примечание"]
        for i, field in enumerate(fields, start=1):
            tk.Label(right_frame, text=field).pack(anchor="w")
            entry_var = tk.StringVar()
            entry = tk.Entry(right_frame, textvariable=entry_var, width=60)
            entry.pack(anchor="w", pady=5)

            self.dynamic_fields[field] = entry_var

        # Load Excel headers and create fields in right_frame
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook.worksheets[1]  # Access the second sheet
            headers = [cell.value for cell in sheet[1]]

            # Skip the first 5 headers and start from column 6
            for i, header in enumerate(headers[5:], start=6):
                tk.Label(right_frame, text=header).pack(anchor="w")
                entry_var = tk.StringVar()
                entry = tk.Entry(right_frame, textvariable=entry_var, width=60)
                entry.pack(anchor="w", pady=5)

                self.dynamic_fields[header] = entry_var

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при чтении второго листа: {e}")

        self.update_combobox()

    def add_to_second_sheet(self):
            try:
                workbook = load_workbook(self.filepath)
                sheet = workbook.worksheets[1]
                
                tome_number = self.dynamic_fields["Номер тома"].get()
                if tome_number:
                    existing_tome_numbers = [cell.value for cell in sheet["A"] if cell.value]
                    if tome_number in existing_tome_numbers:
                        messagebox.showerror("Ошибка", "Том с таким номером уже существует")
                        return

                for row in range(sheet.max_row + 1, sheet.max_row + 1000):
                    if all(cell.value is None for cell in sheet[row]):
                        break
                else:
                    messagebox.showerror("Ошибка", "Второй лист переполнен")
                    return

                for i, (header, entry_var) in enumerate(self.dynamic_fields.items(), start=1):
                    if i <= 4:
                        sheet.cell(row=row, column=i).value = entry_var.get()
                    else:
                        sheet.cell(row=row, column=i+1).value = entry_var.get()

                workbook.save(self.filepath)
                messagebox.showinfo("Успех", "Данные успешно сохранены в Excel файл.")

                self.update_combobox()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при записи в файл: {e}")

    def delete_selected_item(self):
        selected_value = self.dynamic_combobox.get()

        if selected_value:
            tome_number = selected_value.split(" - ")[0]
            try:
                workbook = load_workbook(self.filepath)
                sheet = workbook.worksheets[1]

                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == str(tome_number):
                        sheet.delete_rows(row)
                        break
                else:
                    messagebox.showerror("Ошибка", f"Номер тома {tome_number} не найден")
                workbook.save(self.filepath)
                messagebox.showinfo("Успех", "Выбранная строка успешно удалена")

                self.update_combobox()
                self.clear_fields()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при удалении строки: {e}")
    def edit_selected_item(self):
        selected_value = self.dynamic_combobox.get()

        if selected_value:
            tome_number = selected_value.split(" - ")[0]
            try:
                workbook = load_workbook(self.filepath)
                sheet = workbook.worksheets[1]
                tome_number2 = self.dynamic_fields["Номер тома"].get()
                if tome_number2:
                    existing_tome_numbers = [cell.value for cell in sheet["A"] if cell.value]
                    if existing_tome_numbers.count(tome_number2) > 1:
                        messagebox.showerror("Ошибка", "Том с таким номером уже существует")
                        return
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == tome_number:
                        break
                else:
                    messagebox.showerror("Ошибка", "Выбранный том не найден в файле")
                    return

                for i, (header, entry_var) in enumerate(self.dynamic_fields.items(), start=1):
                    if i <= 4:
                        sheet.cell(row=row, column=i).value = entry_var.get()
                    else:
                        sheet.cell(row=row, column=i+1).value = entry_var.get()

                workbook.save(self.filepath)
                messagebox.showinfo("Успех", "Данные успешно изменены в Excel файле.")

                self.update_combobox()
                self.dynamic_combobox.set(selected_value)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при изменении файла: {e}")

    def clear_fields(self):
        # Clear all entry fields
        for var in self.dynamic_fields.values():
            var.set("")
        
        # Reset the combobox selection
        self.dynamic_combobox.set("")

    def clear_tab(self, tab):
        for widget in tab.winfo_children():
            widget.destroy()
    def load_image_path(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg"), ("Image files", "*.png")])
        if file_path:
            self.image_path.set(file_path)
            self.load_image()
    def load_pdf_file(self):
        selected_value = self.dynamic_combobox.get()
        if not selected_value:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите элемент из списка.")
            return

        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if file_path:
            tome_number = selected_value.split(" - ")[0]
            try:
                workbook = load_workbook(self.filepath)
                sheet = workbook.worksheets[1]

                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == tome_number:
                        sheet.cell(row=row, column=5).value = file_path
                        break

                workbook.save(self.filepath)
                messagebox.showinfo("Успех", "Путь к PDF файлу успешно сохранён в Excel файл.")

                # Update the PDF path label with the new file path
                self.pdf_path_label.config(text=f"Путь к pdf файлу: {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при записи файла: {e}")

    def select_file(self):
        self.filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            self.load_excel_data()
            self.create_project_info_page()
            self.create_project_contents_page()
            self.notebook.forget(self.tab_file_selection)
            self.notebook.add(self.tab_project_info, text="Информация о проекте")
            self.notebook.add(self.tab_project_contents, text="Состав проекта")

            
        try:
            workbook = load_workbook(self.filepath)
            if "Сведения о проекте" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Сведения о проекте")
                sheet.append(["Заказчик", "Наименование проекта", "Фон", "Шифр проекта", "Генеральный директор", 
                              "Главный инженер проекта", "Город", "Год разработки"])
            if "Состав проекта" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Состав проекта")
                sheet.append(["Номер тома", "Обозначение", "Наименование", "Примечание", "pdf-файл", "Разработал", "Проверил", "ГИП", "Н.Контроль"])
            if workbook.sheetnames[0] != "Сведения о проекте" and workbook.sheetnames[0] != "Состав проекта":
                del workbook[workbook.sheetnames[0]]

                
                workbook.save(self.filepath)
        except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при создании листов: {e}")

    def load_excel_data(self):
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook.worksheets[0]  # Access the first sheet
            self.customer.set(sheet["A2"].value or "")
            self.project_name.set(sheet["B2"].value or "")

            self.image_path.set(sheet["C2"].value or "")
            self.project_code.set(sheet["D2"].value or "")
            self.ceo.set(sheet["E2"].value or "")
            self.project_engineer.set(sheet["F2"].value or "")
            self.city.set(sheet["G2"].value or "")
            self.year_of_development.set(sheet["H2"].value or "")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при чтении файла: {e}")
            print(e)

    def load_image(self):
        try:
            if self.image_path.get():
                img = Image.open(self.image_path.get())
                img = img.resize((465, 657), Image.LANCZOS)
                img = ImageTk.PhotoImage(img)
                self.image_panel.config(image=img)
                self.image_panel.image = img
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке изображения: {e}")
    def overlay_text_on_image(self):
        output_path_docx = "temptitle.docx"
        output_path_pdf = "temptitle.pdf"
        output_path_png = "temptitle.png"

        try:
            doc = DocxTemplate("титул_шаблон_пример.docx")

            replacements = {
                'Заказчик': self.customer.get(),
                'Наименование': self.project_name.get(),
                'Шифр': self.project_code.get(),
                'Директор': self.ceo.get(),
                'Инженер': self.project_engineer.get(),
                'Город': self.city.get(),
                'Год': self.year_of_development.get()
            }

            doc.render(replacements)
            doc.save(output_path_docx)

            convert(output_path_docx, output_path_pdf)

            reader = PdfReader(output_path_pdf)
            writer = PdfWriter()
            page = reader.pages[0]

            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height)

            if self.image_path.get():
                packet = BytesIO()
                c = canvas.Canvas(packet, pagesize=(page_width, page_height))
                c.drawImage(self.image_path.get(), 0, 0, width=page_width, height=page_height)
                c.save()
                packet.seek(0)

                background_pdf = PdfReader(packet)
                background_page = background_pdf.pages[0]
                background_page.merge_page(page)
                writer.add_page(background_page)
            else:
                writer.add_page(page)

            with open(output_path_pdf, "wb") as f:
                writer.write(f)

            pdf_document = fitz.open(output_path_pdf)
            for page_number in range(len(pdf_document)):
                page = pdf_document.load_page(page_number)
                pix = page.get_pixmap()
                pix.save(output_path_png)

            img = Image.open(output_path_png)
            img = img.resize((465, 657), Image.LANCZOS)
            img = ImageTk.PhotoImage(img)
            self.image_panel.config(image=img)
            self.image_panel.image = img

        finally:
            for temp_file in [output_path_docx, output_path_pdf, output_path_png]:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    print(f"Не удалось удалить {temp_file}: {e}")


    def populate_fields_from_selection(self, event):
        
        selected_value = self.dynamic_combobox.get()
        
        if selected_value:
            tome_number = selected_value.split(" - ")[0]
            try:
                workbook = load_workbook(self.filepath)
                sheet = workbook.worksheets[1]

                # Iterate over rows to find the selected "Номер тома"
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == tome_number:
                        # Populate dynamic fields based on the row data
                        for i, (header, entry_var) in enumerate(self.dynamic_fields.items()):
                            if i >= 4:
                                i += 1
                            entry_var.set(row[i] if i < len(row) else "")

                        # Update the pdf_path_label text with the value from column E (PDF path)
                        self.pdf_path_label.config(text=f"Путь к pdf файлу: {row[4]}")
                        break
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при заполнении полей: {e}")
 
    def save_to_excel(self):
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook.worksheets[0]  # Access the first sheet
            sheet["A2"] = self.customer.get()
            sheet["B2"] = self.project_name.get()
            sheet["C2"] = self.image_path.get()
            sheet["D2"] = self.project_code.get()  # Save new field
            sheet["E2"] = self.ceo.get()  # Save new field
            sheet["F2"] = self.project_engineer.get()  # Save new field
            sheet["G2"] = self.city.get()  # Save new field
            sheet["H2"] = self.year_of_development.get()  # Save new field

            workbook.save(self.filepath)
            messagebox.showinfo("Успех", "Данные успешно сохранены в Excel файл.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при сохранении файла: {e}")

    def update_combobox(self):
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook.worksheets[1]  # Access the second sheet
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:  # Ensure there's data in these columns
                    rows.append(row)
            rows.sort(key=lambda row: row[0])
            combobox_values = [f"{row[0]} - {row[2]}" for row in rows]

            # Update the Combobox with the collected values
            self.dynamic_combobox["values"] = combobox_values
            if combobox_values:
                self.dynamic_combobox.current(0)  # Set to first item if there are values
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обновлении списка: {e}")

    def create_word_documents(self):
        pass
root = tk.Tk()
app = ProjectInfoApp(root)
root.mainloop()

